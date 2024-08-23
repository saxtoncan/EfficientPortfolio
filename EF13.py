import matplotlib
matplotlib.use('Agg')  # Use the 'Agg' backend for non-GUI operations
import matplotlib.pyplot as plt
import yfinance as yf
import pandas as pd
import numpy as np
import scipy.optimize as opt
import tkinter as tk
from tkinter import messagebox, simpledialog
from tkinter import ttk
from tkcalendar import DateEntry
from threading import Thread
from yahooquery import search
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import io  # Import io for in-memory byte streams

# Global flag to stop fetching process
cancel_flag = False

# Function to fetch data for a single ticker
def fetch_stock_data(ticker, start_date, end_date, interval):
    try:
        print(f"Fetching data for {ticker} with interval {interval}")
        stock_data = yf.download(ticker, start=start_date, end=end_date, interval=interval)

        if stock_data.empty:
            print(f"Data for {ticker} is unavailable or returned empty.")
            return ticker, None, None

        # Ensure the index is datetime
        stock_data.index = pd.to_datetime(stock_data.index)

        # Calculate log returns using 'Adj Close' prices and store them
        stock_data['Log Returns'] = stock_data['Adj Close'].apply(lambda x: np.log(x)).diff()
        log_returns_data = stock_data.dropna(subset=['Log Returns'])
        return ticker, stock_data, log_returns_data

    except Exception as e:
        error_message = f"Error fetching data for {ticker}: {e}"
        print(error_message)
        return ticker, None, None

# Function to fetch data and save to Excel
def fetch_and_save_data(tickers, start_date, end_date, interval, filename, progress_var, progress_label):
    global cancel_flag
    total_tickers = len(tickers)

    # Create DataFrames to store combined stock data and log returns
    combined_data = {stat: pd.DataFrame() for stat in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']}
    combined_log_returns = pd.DataFrame()

    # Determine the annualization factor based on the interval
    if interval == '1d':
        annualization_factor = 252
    elif interval == '1wk':
        annualization_factor = 52
    elif interval == '1mo':
        annualization_factor = 12
    else:
        annualization_factor = 252  # Default to daily if unknown

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for i, ticker in enumerate(tickers):
            if cancel_flag:
                messagebox.showinfo("Cancelled", "Data fetching process was cancelled.")
                break

            # Fetch data for the ticker
            ticker, stock_data, log_returns_data = fetch_stock_data(ticker, start_date, end_date, interval)

            if stock_data is not None:
                # Combine individual data for each statistic
                for stat in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']:
                    combined_data[stat][ticker] = stock_data[stat]

                # Combine log returns for each stock
                combined_log_returns[ticker] = log_returns_data['Log Returns']

            # Update progress
            progress = int((i + 1) / total_tickers * 100)
            progress_var.set(progress)
            progress_label.config(text=f'Progress: {progress}%')

        # Write combined data for each statistic to separate sheets
        for stat in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']:
            combined_data[stat].to_excel(writer, sheet_name=stat)

        # Write combined log returns to a new sheet, without the first row (blank log returns)
        if not combined_log_returns.empty:
            combined_log_returns = combined_log_returns.dropna()
        combined_log_returns.to_excel(writer, sheet_name='Log Returns')

        # Calculate and write efficient portfolio and efficient frontier
        if not combined_log_returns.empty:
            calculate_and_write_efficient_portfolio_and_frontier(combined_log_returns, writer, annualization_factor)

        # Calculate and write correlation matrix
        if not combined_log_returns.empty:
            correlation_matrix = combined_log_returns.corr()
            correlation_matrix.to_excel(writer, sheet_name='Correlation Matrix')

        # Calculate and write covariance matrix
        if not combined_log_returns.empty:
            cov_matrix = combined_log_returns.cov() * annualization_factor
            cov_matrix.to_excel(writer, sheet_name='Covariance Matrix')

    # Re-open the file to adjust date format and column width
    wb = load_workbook(filename)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # Apply General format to column A in Efficient Frontier sheet
        if sheetname == 'Efficient Frontier':
            for cell in ws['A']:
                cell.number_format = 'General'
        else:
            for cell in ws['A']:
                if cell.row == 1:
                    continue
                cell.number_format = 'yyyy-mm-dd'
        ws.column_dimensions['A'].width = 70 / 7  # Set column width to 70 pixels

    wb.save(filename)

    if not cancel_flag:
        messagebox.showinfo("Success", "Data saved successfully to " + filename)

    # Reset cancel flag
    cancel_flag = False

# Function to calculate and write the efficient portfolio and efficient frontier
def calculate_and_write_efficient_portfolio_and_frontier(log_returns, writer, annualization_factor):
    if log_returns.empty or len(log_returns.columns) < 2:
        print("Not enough data to calculate the efficient frontier.")
        return

    returns = log_returns.mean() * annualization_factor
    cov_matrix = log_returns.cov() * annualization_factor
    num_assets = len(returns)
    num_portfolios = 10000

    min_weight = float(min_weight_var.get())
    max_weight = float(max_weight_var.get())
    min_weights = [min_weight] * num_assets
    max_weights = [max_weight] * num_assets

    results = np.zeros((3, num_portfolios))
    all_weights = np.zeros((num_portfolios, num_assets))

    for i in range(num_portfolios):
        weights = np.random.random(num_assets)
        weights /= np.sum(weights)

        portfolio_return = np.sum(weights * returns)
        portfolio_volatility = np.sqrt(np.dot(weights.T, np.dot(cov_matrix, weights)))
        portfolio_sharpe = portfolio_return / portfolio_volatility

        results[0, i] = portfolio_return
        results[1, i] = portfolio_volatility
        results[2, i] = portfolio_sharpe
        all_weights[i, :] = weights

    results_df = pd.DataFrame(results.T, columns=['Return', 'Volatility', 'Sharpe Ratio'])
    results_df.to_excel(writer, sheet_name='Efficient Frontier', index=False)

    # Optimal portfolio
    def portfolio_performance(weights):
        p_return = np.sum(weights * returns)
        p_volatility = np.sqrt(np.dot(weights.T, np.dot(cov_matrix, weights)))
        return p_return, p_volatility

    def min_sharpe_ratio(weights):
        return -portfolio_performance(weights)[0] / portfolio_performance(weights)[1]

    def check_sum(weights):
        return np.sum(weights) - 1

    constraints = ({'type': 'eq', 'fun': check_sum})
    bounds = tuple((min_weight, max_weight) for _ in range(num_assets))
    initial_guess = np.array(num_assets * [1. / num_assets])

    # Run the optimizer
    opt_result = opt.minimize(min_sharpe_ratio, initial_guess, method='SLSQP', bounds=bounds, constraints=constraints)
    optimal_weights = opt_result.x
    optimal_return, optimal_volatility = portfolio_performance(optimal_weights)

    optimal_df = pd.DataFrame({
        'Stock': log_returns.columns,
        'Weight': optimal_weights
    })

    # Store the weights as numeric values and format them as percentages with three decimal places in Excel
    optimal_df['Weight'] = optimal_df['Weight'].round(5)  # Round to 5 decimal places for precision

    # Add the expected return of the optimal portfolio
    expected_return = np.sum(optimal_weights * returns)
    expected_return_df = pd.DataFrame([{'Stock': 'Expected Return', 'Weight': expected_return}])

    optimal_df = pd.concat([optimal_df, expected_return_df], ignore_index=True)
    optimal_df.to_excel(writer, sheet_name='Optimal Portfolio', index=False)

    # Adjust the formatting of the 'Weight' column to show percentages with three decimal places
    workbook = writer.book
    worksheet = writer.sheets['Optimal Portfolio']

    for idx, _ in enumerate(optimal_df.index):
        worksheet[f'B{idx + 2}'].number_format = '0.000%'  # Applies percentage format with three decimal places

    # Apply conditional formatting to highlight cells with weights not equal to zero
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    worksheet.conditional_formatting.add(
        f'B2:B{len(optimal_df) + 1}',
        CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=yellow_fill)
    )

    # Retrieve the annual risk-free rate from the GUI and convert it to the appropriate interval
    interval = interval_var.get()
    annual_risk_free_rate = float(risk_free_rate_var.get()) / 100  # Convert to decimal

    # Scale the risk-free rate based on the selected interval
    if interval == '1d':
        scaled_risk_free_rate = (1 + annual_risk_free_rate) ** (1/252) - 1  # Daily data
    elif interval == '1wk':
        scaled_risk_free_rate = (1 + annual_risk_free_rate) ** (1/52) - 1  # Weekly data
    elif interval == '1mo':
        scaled_risk_free_rate = (1 + annual_risk_free_rate) ** (1/12) - 1  # Monthly data
    else:
        scaled_risk_free_rate = annual_risk_free_rate  # Default to no scaling if interval is unknown

    # Find the portfolio with the maximum Sharpe ratio (Tangency Portfolio)
    max_sharpe_idx = np.argmax(results_df['Sharpe Ratio'])
    max_sharpe_return = results_df['Return'][max_sharpe_idx]
    max_sharpe_volatility = results_df['Volatility'][max_sharpe_idx]

    # Plot the Efficient Frontier
    plt.figure(figsize=(10, 6))

    # Plot the scatter plot for the Efficient Frontier
    scatter = plt.scatter(results_df['Volatility'], results_df['Return'], c=results_df['Sharpe Ratio'], cmap='viridis',
                          alpha=0.5)
    plt.colorbar(scatter, label='Sharpe Ratio')

    # Plot the Tangency Portfolio
    plt.scatter(max_sharpe_volatility, max_sharpe_return, marker='*', color='r', s=200, label='Tangency Portfolio')

    # Plot the Capital Allocation Line (CAL)
    cal_x = np.linspace(0, results_df['Volatility'].max(), 100)
    cal_y = scaled_risk_free_rate + (max_sharpe_return - scaled_risk_free_rate) / max_sharpe_volatility * cal_x
    plt.plot(cal_x, cal_y, linestyle='-', color='orange', label='Capital Allocation Line (CAL)')

    plt.xlabel('Volatility (Risk)')
    plt.ylabel('Return')
    plt.title('Efficient Frontier with CAL and Tangency Portfolio')
    plt.legend(loc='best')
    plt.grid(True)

    # Save the plot to a BytesIO object and insert into Excel
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png')
    plt.close()
    img_buffer.seek(0)

    # Insert the image into the Excel sheet
    worksheet_frontier = writer.sheets['Efficient Frontier']
    img = Image(img_buffer)
    worksheet_frontier.add_image(img, 'E2')  # Adjust the cell position as needed

# Function to handle the fetch button click
def on_fetch_click():
    global cancel_flag
    cancel_flag = False
    tickers = [ticker.strip() for ticker in ticker_entry.get().split(',')]
    start_date = start_date_entry.get_date().strftime('%Y-%m-%d')
    end_date = end_date_entry.get_date().strftime('%Y-%m-%d')
    interval = interval_var.get()
    filename = simpledialog.askstring("Input", "Enter the filename for the Excel file:",
                                      initialvalue="Efficient_Portfolio.xlsx")
    if filename:
        # Run data fetching in a separate thread to keep the GUI responsive
        thread = Thread(target=fetch_and_save_data,
                        args=(tickers, start_date, end_date, interval, filename, progress_var, progress_label))
        thread.start()

# Function to handle the cancel button click
def on_cancel_click():
    global cancel_flag
    cancel_flag = True

# Function to search for ticker symbol and insert it into the ticker entry
def search_ticker():
    query = search_entry.get().strip()
    if query:
        try:
            results = search(query)
            if results and 'quotes' in results and results['quotes']:
                matches = [(res['symbol'], res.get('shortname', 'No Name')) for res in results['quotes']]

                if matches:
                    result_window = tk.Toplevel(root)
                    result_window.title("Search Results")

                    tk.Label(result_window, text="Select a ticker symbol:").pack(pady=10)

                    ticker_var = tk.StringVar()

                    # Create a listbox for displaying matches
                    result_listbox = tk.Listbox(result_window, width=50)
                    result_listbox.pack(pady=10)

                    # Add matches to the listbox
                    for symbol, name in matches:
                        result_listbox.insert(tk.END, f"{symbol}: {name}")

                    def on_select():
                        selected_indices = result_listbox.curselection()
                        if selected_indices:
                            selected = result_listbox.get(selected_indices[0])
                            ticker_symbol = selected.split(":")[0].strip()
                            # Avoid adding duplicate commas
                            existing = ticker_entry.get()
                            if existing and not existing.endswith(','):
                                ticker_entry.insert(tk.END, f", {ticker_symbol}")
                            else:
                                ticker_entry.insert(tk.END, f"{ticker_symbol}")
                            result_window.destroy()

                    select_button = tk.Button(result_window, text="Select", command=on_select)
                    select_button.pack(pady=10)

        except Exception as e:
            print(f"Error searching ticker: {e}")

# Create the main window
root = tk.Tk()
root.title("Efficient Portfolio")

# Ticker entry
tk.Label(root, text="Enter ticker symbols (comma separated):").grid(row=0, column=0, pady=10, padx=10, sticky='e')
ticker_entry = tk.Entry(root, width=50)
ticker_entry.grid(row=0, column=1, pady=10, padx=10)

# Search ticker entry and button
tk.Label(root, text="Search for a ticker:").grid(row=1, column=0, pady=10, padx=10, sticky='e')
search_entry = tk.Entry(root, width=50)
search_entry.grid(row=1, column=1, pady=10, padx=10)
search_button = tk.Button(root, text="Search", command=search_ticker)
search_button.grid(row=1, column=2, pady=10, padx=10)

# Date range selection
tk.Label(root, text="Start date:").grid(row=2, column=0, pady=10, padx=10, sticky='e')
start_date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2)
start_date_entry.grid(row=2, column=1, pady=10, padx=10, sticky='w')

tk.Label(root, text="End date:").grid(row=3, column=0, pady=10, padx=10, sticky='e')
end_date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2)
end_date_entry.grid(row=3, column=1, pady=10, padx=10, sticky='w')

# Interval selection
tk.Label(root, text="Select interval:").grid(row=4, column=0, pady=10, padx=10, sticky='e')
interval_var = tk.StringVar(value='1d')
ttk.Combobox(root, textvariable=interval_var, values=['1d', '1wk', '1mo'], width=10).grid(row=4, column=1, pady=10, padx=10, sticky='w')

# Weight constraints
tk.Label(root, text="Minimum weight:").grid(row=5, column=0, pady=10, padx=10, sticky='e')
min_weight_var = tk.StringVar(value='0')
tk.Entry(root, textvariable=min_weight_var, width=5).grid(row=5, column=1, pady=10, padx=10, sticky='w')

tk.Label(root, text="Maximum weight:").grid(row=6, column=0, pady=10, padx=10, sticky='e')
max_weight_var = tk.StringVar(value='1')
tk.Entry(root, textvariable=max_weight_var, width=5).grid(row=6, column=1, pady=10, padx=10, sticky='w')

# Risk-free rate input
tk.Label(root, text="Risk-free rate (%):").grid(row=7, column=0, pady=10, padx=10, sticky='e')
risk_free_rate_var = tk.StringVar(value='2')  # Default value of 2%
tk.Entry(root, textvariable=risk_free_rate_var, width=5).grid(row=7, column=1, pady=10, padx=10, sticky='w')

# Note for the risk-free rate input
tk.Label(root, text="* Enter the annual risk-free rate. It will be adjusted based on the selected interval.",
         fg="blue").grid(row=8, column=0, columnspan=2, pady=5, padx=10)

# Progress bar and label
progress_var = tk.IntVar(value=0)
progress_bar = ttk.Progressbar(root, length=400, variable=progress_var)
progress_bar.grid(row=9, column=0, columnspan=3, pady=20)
progress_label = tk.Label(root, text="Progress: 0%")
progress_label.grid(row=10, column=0, columnspan=3)

# Fetch and cancel buttons
fetch_button = tk.Button(root, text="Fetch Data", command=on_fetch_click)
fetch_button.grid(row=11, column=0, pady=20)

cancel_button = tk.Button(root, text="Cancel", command=on_cancel_click)
cancel_button.grid(row=11, column=1, pady=20)

root.mainloop()